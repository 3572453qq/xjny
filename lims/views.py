from django.shortcuts import render
import random,string
import pandas as pd
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import re
import shutil
import os
import random
import string
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side
import math
from openpyxl.utils import get_column_letter
import base64
import pymysql
import multiprocessing
from lims.models import *

border_style = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

def copy_excel_to_tmp(src_file_path):
    # 获取原始文件名和扩展名
    file_name, file_extension = os.path.splitext(os.path.basename(src_file_path))

    # 生成7位随机数字
    random_number = ''.join(random.choices(string.digits, k=7))

    # 获取当前日期并格式化为YYYYMMDD
    current_date = datetime.now().strftime('%Y%m%d')

    # 构建目标文件名
    dest_file_name = f"{current_date}_result_{random_number}{file_extension}"


    # 构建目标文件路径
    dest_file_path = os.path.join('/tmp', dest_file_name)

    # 拷贝文件到目标路径
    shutil.copy(src_file_path, dest_file_path)

    return dest_file_path

# Create your views here.
def base(request):
    return render(request,'lims/base.html')

def base_test(request):
    return render(request,'lims/base_test.html')

def vkeepexcel(request):
    return render(request,'lims/vkeepexcel.html')

def querycycle(request):
    context_dict = {'module': 'lims'}
    all_teams = list(teams.objects.values())
    context_dict['teams']=all_teams
    return render(request,'lims/querycycle.html',context_dict)

def handlevkeep(request):
    if request.method == 'POST' and request.FILES.getlist('files'):
        response_data = []
        summary_data = []
        print('hello')
        print( request.FILES.getlist('files'))
        
        max_length = 0
        for afile in request.FILES.getlist('files'): 
            df_result = pd.read_excel(afile,sheet_name='cycle')  
            current_length = df_result.shape[0]
            if max_length < current_length:
                max_length=current_length
        chart_df = pd.DataFrame(index=range(max_length))
        # 处理每个上传的文件
        for filename in request.FILES.getlist('files'):
            # print(filename)
            # 使用 pandas 读取 Excel 文件
            df = pd.read_excel(filename,sheet_name='cycle')
            df_result = df[['循环号','充电容量(Ah)','放电容量(Ah)']]
            df_result['保持率'] = (df_result['放电容量(Ah)'] / df_result['放电容量(Ah)'].iloc[0] ).round(4)
            df_result['库伦效率'] = (df_result['放电容量(Ah)'] / df_result['充电容量(Ah)'] ).round(4)



            # print(df_result)
            # 获取列头信息
            columns = list(df_result.columns)
            # 去掉字符串里面的括号，否则会导致kendogrid生成不了
            columns =[re.sub(r'\(|\)|\%','',item) for item in columns]
            # print(columns)
            # 将 DataFrame 转换为 JSON 格式
            json_data = df_result.to_dict(orient='records')
            
            json_data = [{key.replace('(', '').replace(')', '').replace('%', ''): value for key,value in d.items()} for d in json_data]
            # print(json_data)
            json_data = json.dumps(json_data)

            barcode = re.sub(r'\.[^.]*$','', filename.name)
            chart_df[barcode] = df_result['保持率']
            # 将数据和文件名添加到响应数据中
            response_data.append({
                'filename': filename.name,
                'barcode': barcode,
                'data': json_data,
                'columns': columns
            })

            summary_data.append({
                '电芯编码': barcode,
                '通道': '待填写',
                '分容容量Ah':'待填写',
                '放电电流':'待填写',
                '测试后电压':'待填写',
                '测试后内阻':'待填写',
                '测试后厚度':'待填写',

            })

        
        summary_data = json.dumps(summary_data)
        
        # 单独抽出来的保持率搞一个列表
        chart_df = chart_df.fillna(0)
        # 转换为二维列表
        two_dimensional_list = chart_df.values.tolist()
        # 获取列头
        column_headers = list(chart_df.columns)
        # 在二维列表中插入列头作为第一行
        two_dimensional_list.insert(0, column_headers)
        
        # print('*'*1000)
        return JsonResponse({'data': response_data,'summary':summary_data,'chart':two_dimensional_list}, status=200)
    return JsonResponse({'error': 'Invalid request'}, status=400)


def cyclesort(request):
    return render(request,'lims/cyclesort.html')

def cycles(request):
    return render(request,'lims/cycles.html')

def crate(request):
    return render(request,'lims/crate.html')


def write_dataframes_to_excel(input_excel_path, dataframes_with_titles,summary_result):
    # Load existing workbook
    wb = load_workbook(input_excel_path)
    ws = wb.active

    # Determine the starting column (17th column)
    start_column = 17

    # 创建一个百分比样式
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    wb.add_named_style(percent_style)



    # 写入summary数据区域
    for rownum,row in enumerate(summary_result, start=2):
        print(row,rownum)
        ws.cell(row=rownum,column=1,value=row[0])
        ws.cell(row=rownum,column=1).border = border_style
        ws.cell(row=rownum,column=11,value=row[1])
        ws.cell(row=rownum,column=11).border = border_style
        ws.cell(row=rownum,column=12,value=row[2])
        ws.cell(row=rownum,column=12).border = border_style
        ws.cell(row=rownum,column=13,value=row[3])
        ws.cell(row=rownum,column=13).border = border_style 
        
    

    # Iterate over the dataframes and titles
    for df, title in dataframes_with_titles:
        # print(df)
        df = df.drop(columns=['循环号'])
        print(df)
        # Write title
        ws.cell(row=1, column=start_column, value=title)
        start_cell = get_column_letter(start_column) + '1'
        end_cell = get_column_letter(start_column+3) + '1'
        print(start_cell,end_cell)
        ws.merge_cells(start_cell + ':' + end_cell)
        ws.cell(row=1,column=start_column).border = border_style

        # Write column names
        # for col_num, col_name in enumerate(df.columns, start=start_column):
        #     ws.cell(row=2, column=col_num, value=col_name)

        # Write data
        for row_num, row in enumerate(dataframe_to_rows(df, index=False), start=2):
            # print(row,row_num)
            for col_num, value in enumerate(row, start=start_column):
                if value is None:
                    break
                ws.cell(row=row_num, column=col_num, value=value)
                if (col_num-start_column==2 or col_num-start_column==3) and row_num>2:
                    ws.cell(row=row_num,column=col_num).style='percent_style'
                # 将边框样式应用到单元格
                ws.cell(row=row_num,column=col_num).border = border_style
        

        # Update starting column for the next DataFrame
        start_column += len(df.columns)   # +1 for the title

    # Save the updated workbook
    wb.save(input_excel_path)


def handlecycle(request):
    copied_file_path = copy_excel_to_tmp('/data/django/xjny/lims/测试模板.xlsx')

    if request.method == 'POST' and request.FILES.getlist('files'):
        response_data = []
 
        afile = request.FILES.getlist('files')[0]
        df_result = pd.read_excel(afile,sheet_name='Cycle',header=[0, 1])  
        # print(df_result)
        
        # 获取列头一的所有不同值
        column_level_0_values = df_result.columns.get_level_values(0).unique()

        # 遍历不同的列头一，获取每个子 DataFrame
        sub_dataframes = {}
        for level_0_value in column_level_0_values:
            # sub_df = df_result[df_result[(level_0_value, 'Unnamed: 0_level_1')] == level_0_value]
            sub_df = df_result[level_0_value]
            sub_dataframes[level_0_value] = sub_df


        dataframes_with_titles = []
        summary_result = []
        # 输出子 DataFrame
        for key, value in sub_dataframes.items():
            # print(f"Sub DataFrame for {key}:\n{value}\n")
            
            df_sub_result = value[['循环号','充电容量(Ah)','放电容量(Ah)']]
            # 去掉nan的，以及第一行和最后一行
            df_sub_result = df_sub_result.dropna(how='all')
            df_sub_result = df_sub_result.drop([0, df_sub_result.index[-1]])

            df_sub_result['保持率'] = (df_sub_result['放电容量(Ah)'] / df_sub_result['放电容量(Ah)'].iloc[0] ).round(4)
            df_sub_result['库伦效率'] = (df_sub_result['放电容量(Ah)'] / df_sub_result['充电容量(Ah)'] ).round(4)
            
            
            barcode = key
            last_cycle = df_sub_result.loc[df_sub_result.index[-1], '循环号'] - 1
            last_capa_keep = df_sub_result.loc[df_sub_result.index[-1], '保持率']
            last_kulun = df_sub_result.loc[df_sub_result.index[-1], '库伦效率']
            
            dataframes_with_titles.append((df_sub_result,key))
            summary_result.append([barcode,last_cycle,last_capa_keep,last_kulun])
        
        write_dataframes_to_excel(copied_file_path, dataframes_with_titles,summary_result)

        print(copied_file_path)
        # Encode the processed file data as base64
        with open(copied_file_path, 'rb') as file:
            encoded_data = base64.b64encode(file.read()).decode('utf-8')

        # Provide the processed file data and name in the response
        response_data = {
            'file_data': encoded_data,
            'file_name': 'result.xlsx'
        }
     

        return JsonResponse(response_data)
    return JsonResponse({'error': 'Invalid request'}, status=400)

def crate_sheet(result_file,df,sheet_name):
    print(result_file)
    grouped_df = df.groupby('工步号')

    if os.path.isfile(result_file):
        writer = pd.ExcelWriter(result_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') 
    else:
        writer = pd.ExcelWriter(result_file, engine='xlsxwriter') 
       
    start_col = 0        
    for group_name, group_df in grouped_df:
        print(f"工步号: {group_name}")
        print(group_df)
        print("-------------------")
        group_df.to_excel(writer, sheet_name=sheet_name, startcol=start_col, index=False)
        start_col = len(group_df.columns)+start_col
    writer.close()


          


def handlecrate(request):  

    if request.method == 'POST' and request.FILES.getlist('files'):
        response_data = []

        # for afile in  request.FILES.getlist('files'):
        afile = request.FILES.getlist('files')[0]
        # 生成5位随机数字
        random_number = ''.join(random.choices(string.digits, k=5))

        result_file = '/tmp/r'+random_number+afile.name 
        df_record = pd.read_excel(afile,sheet_name='record',header=[0]) 

        df_disch = df_record[df_record['工步类型'] == '恒流放电'][['工步号','电压(V)','容量(Ah)']]
        crate_sheet(result_file,df_disch,'恒流放电') 

        df_chg = df_record[df_record['工步类型'] == '恒流恒压充电'] [['工步号','电压(V)','容量(Ah)']]
        crate_sheet(result_file,df_chg,'恒流恒压充电') 
            
        with open(result_file, 'rb') as file:
                encoded_data = base64.b64encode(file.read()).decode('utf-8')

            # Provide the processed file data and name in the response
        response_data = {
            'file_data': encoded_data,
            'file_name': 'result_'+afile.name
        }
        

        return JsonResponse(response_data)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

def query_database(host,user,password,db,query):
    # 连接到MySQL数据库
    connection = pymysql.connect(host=host, user=user, password=password, db=db)
    try:
        with connection.cursor() as cursor:
            # 执行查询语句
            cursor.execute(query)            
            ls_desc = cursor.description
        ls_columns = []
        for i, ls_column_tuple in enumerate(ls_desc):
            ls_column = ls_column_tuple[0]
            if len(ls_column) < 1:
                ls_column = str(i)

            ls_formated_column = re.sub('\W+', '_', ls_column)
            ls_formated_column = 'a' + \
                ls_formated_column if re.match(
                    '^[0-9]', ls_formated_column) else ls_formated_column

            ls_columns.append(ls_formated_column)
        all_lines = []
        for a_line in cursor.fetchall():
            a_data = dict(zip(ls_columns, a_line))
            all_lines.append(a_data)        
        return (all_lines,ls_columns)
    finally:
        connection.close()

def getcycledata1(request):
    start_date = request.POST.get('startdate')
    end_date = request.POST.get('enddate')
    db_name = request.POST.get('dbname')
    team_id = request.POST.get('teamid')
    print(start_date, end_date,db_name)
    is_ok=1
    try:
        db_host, db_user, db_pass = '172.28.20.100','root','Abc_12345'

        con = pymysql.connect(host=db_host, user=db_user,
                                password=db_pass, database=db_name)
        cursor = con.cursor(cursor=pymysql.cursors.Cursor)

        start_table = 'cycle'+start_date
        end_table = 'cycle'+end_date

        select_table_name = f'''SELECT table_name 
            FROM INFORMATION_SCHEMA.TABLES
            where table_name between '{start_table}' and '{end_table}'
            and table_schema='yanfa' '''
        
        cursor.execute(select_table_name)
        # 构建选择数据的sql
        tables = [table[0] for table in cursor.fetchall()]
        print(tables)
        select_cycle_data = f'''SELECT * FROM {(' UNION ALL SELECT * FROM '.join(tables))} 
                                order by dev_unit_chl,test_id,cycle_id;'''
        print(select_cycle_data)
        cursor.execute(select_cycle_data)
        # cursor.execute('select * from cycle20240213 order by dev_unit_chl,test_id,cycle_id')

        ls_desc = cursor.description
        # log_addition(request.user, pagepermission,
        #                 'sql execution:['+ls_sqlstr+'] on ['+ls_connstr+']')  # 记录日志
    except Exception as e:
        is_ok =0 
        ls_data = {"isok": is_ok, "errmsg": str(e)}
        return JsonResponse(ls_data)
    
    ls_columns = []

    # print(ls_desc)

    for i, ls_column_tuple in enumerate(ls_desc):
        ls_column = ls_column_tuple[0]
        if len(ls_column) < 1:
            ls_column = str(i)

        ls_formated_column = re.sub('\W+', '_', ls_column)
        ls_formated_column = 'a' + \
            ls_formated_column if re.match(
                '^[0-9]', ls_formated_column) else ls_formated_column
        # print('column is',ls_formated_column)
        ls_columns.append(ls_formated_column)

    print('here are the columns', ls_columns)

    all_lines = []
    for a_line in cursor.fetchall():
        a_data = dict(zip(ls_columns, a_line))
        all_lines.append(a_data)

    cursor.close()
    con.close()

    ls_data = {'data': all_lines, 'total': len(
        all_lines), 'columns': ls_columns, "isok": is_ok}

    return JsonResponse(ls_data)


def getcycledata(request):       
    start_date = request.POST.get('startdate')
    end_date = request.POST.get('enddate')
    db_name = request.POST.get('dbname')
    team_id = request.POST.get('teamid')
    cycle_count = int(request.POST.get('cyclecount'))
    chg_efficiency = float(request.POST.get('chgefficiency'))/100

    print(start_date, end_date,db_name,team_id,cycle_count,chg_efficiency)
    is_ok=1

    all_computers = list(teamcomputer.objects.filter(
        teamid=team_id).values('computer_name'))
    
    print('this is all computers',all_computers)
    computer_names = [k['computer_name'] for k in all_computers]
    print('this is computer_names',computer_names)

    try:
        db_host, db_user, db_pass = '172.28.20.100','root','Abc_12345'
        con = pymysql.connect(host=db_host, user=db_user,
                                password=db_pass, database=db_name)
        cursor = con.cursor(cursor=pymysql.cursors.Cursor)

        start_table = 'cycle'+start_date
        end_table = 'cycle'+end_date

        select_table_name = f'''SELECT table_name 
            FROM INFORMATION_SCHEMA.TABLES
            where table_name between '{start_table}' and '{end_table}'
            and table_schema='{db_name}' '''
        
        print(select_table_name)
        cursor.execute(select_table_name)
        # 构建选择数据的sql
        tables = [table[0] for table in cursor.fetchall()]
        print('here is the',tables)
        str_test = "','".join(computer_names)
        print(str_test)
        
        where_clause = f''' where upper(computer_name) in ('{"','".join(computer_names)}') '''
        print(where_clause)
        

        queries = ['SELECT * FROM '+table+where_clause for table in tables]
        print(queries)
        pool = multiprocessing.Pool()
        process_results = []

        for query in queries:
            process = pool.apply_async(query_database, (db_host,db_user,db_pass,db_name,query))
            process_results.append((process, query))
    
        pool.close()
        pool.join()
        result = []
        for process, query in process_results:
            print('Query: {}'.format(query))
            result+=process.get()[0]
            ls_columns=process.get()[1]          

    except Exception as e:
        is_ok =0 
        ls_data = {"isok": is_ok, "errmsg": str(e)}
        return JsonResponse(ls_data)
    

    cursor.close()
    con.close()
     
    # print(result)
    print(len(result)) 
    if len(result) == 0:
        ls_data = {"isok": 0, "errmsg": '无记录'}
        return JsonResponse(ls_data)
    
    


    # 将原始字典列表转换成 DataFrame
    df = pd.DataFrame(result)

    # 按test_id和dev_unit_chl进行分组
    grouped = df.groupby(['test_id', 'dev_unit_chl'],as_index=False)

    # 初始化一个空的DataFrame，用于存储符合条件的组
    filtered_df = pd.DataFrame()

    # 遍历每个分组
    for name, group in grouped:
        # print(group['discharge_capacity'] / group['charge_capacity'])
        
        # 检查每个组是否超过20行，并且第18行的efficiency值是否超过80
        group['保持率'] = (group['discharge_capacity'] / group['discharge_capacity'].iloc[0] ).round(4)
        group['库伦效率'] = (group['discharge_capacity'] / group['charge_capacity'] ).round(4)
        # print(name, group)
        if len(group) >= cycle_count:
            first_k_rows = group.iloc[2:cycle_count-1]
            all_greater_than_08 = (first_k_rows['保持率'] > chg_efficiency).all()

            # and group.iloc[cycle_count-1]['保持率'] > chg_efficiency:
            # 如果符合条件，则将这个组添加到新的DataFrame中
            if all_greater_than_08:
                print(name,'满足条件')
                filtered_df = pd.concat([filtered_df, group])

    if(len(filtered_df.to_dict(orient='records')))<1:
        ls_data = {"isok": 0, "errmsg": '无满足条件的记录'}
        return JsonResponse(ls_data)

    print(filtered_df.columns.tolist())

    # 根据某个字段排序
    df_sorted = filtered_df.sort_values(by=['dev_unit_chl','test_id','cycle_id'])

    # 得到所有字段名称
    ls_columns = df_sorted.columns.tolist()
    print('here are all the columns',ls_columns) 



    # 将排序后的 DataFrame 转换回字典列表
    result = df_sorted.to_dict(orient='records')

    print('total count is:',len(result))
    lenthofresult = len(result)
    if(lenthofresult>5000):
        print('in sending file part')
        random_number = ''.join(random.choices(string.digits, k=5))
        result_file = '/tmp/getcycledata'+random_number+'.xlsx'

        # ls_data = {"isok": 0, "errmsg": '数量超过5000，请直接下载excel文件'}
        df_sorted.to_excel(result_file, index=False)

        with open(result_file, 'rb') as file:
            encoded_data = base64.b64encode(file.read()).decode('utf-8')

            # Provide the processed file data and name in the response
        ls_data = {
            'isok': 2,
            "errmsg": f'数量超过5000，共{lenthofresult}行,请直接下载excel文件',
            'file_data': encoded_data,
            'file_name': result_file
        }
        return JsonResponse(ls_data)
    else:
        print('小于5000行')
        ls_data = {'data': result, 'total': len(
            result), 'columns': ls_columns, "isok": is_ok}

        return JsonResponse(ls_data)